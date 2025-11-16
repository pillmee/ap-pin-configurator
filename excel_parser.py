"""
Excel 파일 파싱 모듈
AP의 핀 목록 엑셀 파일(*.xlsx)을 읽어 필요한 컬럼 정보를 추출
"""
import openpyxl
from typing import List, Dict, Tuple


def parse_pin_list(file_path: str) -> Tuple[Dict[str, Dict], Dict[str, List[str]]]:
    """
    엑셀 파일에서 핀 목록 정보를 파싱하여 ball 맵과 signal 맵 생성
    
    Args:
        file_path: 엑셀 파일 경로
        
    Returns:
        Tuple[ball_map, signal_map]
        
        ball_map: Ball Location을 키로 하는 맵
        {
            "ball_location": {
                "signals": [signal_name],  # Function Index 순서대로 정렬 (리스트 인덱스가 아니라 엑셀의 Function Index 순서)
                "default_function": int    # Default Function의 Function Index (엑셀에 기록된 값)
            }
        }
        
        signal_map: Signal Name을 키로 하는 맵
        {
            "signal_name": [ball_location1, ball_location2, ...]
        }
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    # 헤더 행 찾기 (첫 번째 행이라고 가정)
    header_row = 1
    headers = {}
    
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        if cell.value:
            headers[cell.value] = col_idx
    
    # 필요한 컬럼 확인
    required_columns = ["Ball Name", "Ball Location", "Signal Name", "Function Index", "Default Function"]
    for col in required_columns:
        if col not in headers:
            raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {col}")
    
    # 컬럼 인덱스
    ball_name_col = headers["Ball Name"]
    ball_loc_col = headers["Ball Location"]
    signal_name_col = headers["Signal Name"]
    func_index_col = headers["Function Index"]
    default_func_col = headers["Default Function"]
    
    # 임시 데이터 저장용 (Function Index별로 정렬하기 위해)
    # Ball Location별로 그룹화 (형태 #1과 #2 모두 처리)
    ball_groups = {}  # {ball_location: {function_index: signal_name, default_function: int}}
    
    # 현재 Ball Location 추적 (형태 #1 처리용)
    current_ball_location = None
    current_default_function = None
    
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        ball_name = sheet.cell(row_idx, ball_name_col).value
        ball_location_cell = sheet.cell(row_idx, ball_loc_col).value
        signal_name = sheet.cell(row_idx, signal_name_col).value
        function_index_cell = sheet.cell(row_idx, func_index_col).value
        default_function_cell = sheet.cell(row_idx, default_func_col).value
        
        # 형태 #2: Ball Location이 명시적으로 채워진 경우
        if ball_location_cell is not None and str(ball_location_cell).strip():
            current_ball_location = str(ball_location_cell).strip()
            
            # Default Function 읽기 (Function Index 값이어야 함)
            if default_function_cell is not None:
                try:
                    current_default_function = int(float(default_function_cell))
                except (ValueError, TypeError):
                    current_default_function = 0
        
        # 형태 #1: Ball Location이 비어있으면 이전 Ball Location 사용
        # (Ball Name이 있는 행에서만 Ball Location 갱신)
        
        # Signal Name이 있으면 데이터 추가
        if current_ball_location and signal_name:
            signal_name = str(signal_name).strip()
            
            # Function Index를 정수로 변환
            try:
                function_index = int(float(function_index_cell)) if function_index_cell is not None else 0
            except (ValueError, TypeError):
                function_index = 0
            
            # Ball Location 그룹 초기화
            if current_ball_location not in ball_groups:
                ball_groups[current_ball_location] = {
                    "signals": {},
                    "default_function": current_default_function if current_default_function is not None else 0
                }
            
            # Signal 추가
            ball_groups[current_ball_location]["signals"][function_index] = signal_name
    
    # Ball 맵과 Signal 맵 생성
    ball_map = ball_groups
    signal_map = {}
    
    # Signal 맵 구성
    for ball_loc, data in ball_groups.items():
        for func_idx, signal_name in data["signals"].items():
            if signal_name not in signal_map:
                signal_map[signal_name] = []
            if ball_loc not in signal_map[signal_name]:
                signal_map[signal_name].append(ball_loc)
    
    # Ball 맵의 signals를 Function Index 순으로 정렬된 리스트로 변환
    # 리스트의 인덱스 = Function Index, 값 = Signal Name
    for ball_loc, data in ball_map.items():
        # Default Function 유효성 검증 (리스트 변환 전에 수행)
        signal_dict = data["signals"]
        max_func_idx = max(signal_dict.keys()) if signal_dict else 0
        if data["default_function"] < 0 or data["default_function"] > max_func_idx:
            # 유효하지 않으면 0으로 설정
            ball_map[ball_loc]["default_function"] = 0
        
        # 리스트로 변환
        sorted_indices = sorted(signal_dict.keys())
        sorted_signals = [signal_dict[idx] for idx in sorted_indices]
        ball_map[ball_loc]["signals"] = sorted_signals
        # default_function은 Function Index 그대로 유지
    
    workbook.close()
    
    return ball_map, signal_map


def get_default_signal_index(ball_data: Dict) -> int:
    """
    특정 Ball Location의 기본 Signal Function Index 찾기
    
    Args:
        ball_data: ball_map의 값 (특정 ball location의 데이터)
        
    Returns:
        기본 Signal의 Function Index (실제 엑셀의 Function Index 값), 없으면 0
    """
    return ball_data.get("default_function", 0)


if __name__ == "__main__":
    # 테스트 코드
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        try:
            ball_map, signal_map = parse_pin_list(file_path)
            
            print(f"총 {len(ball_map)}개의 Ball Location 발견")
            print(f"총 {len(signal_map)}개의 고유 Signal Name 발견")
            
            # 처음 3개 Ball Location 출력
            print("\n[Ball Map]")
            for idx, (ball_loc, data) in enumerate(list(ball_map.items())[:3]):
                print(f"\nBall Location: {ball_loc}")
                print(f"  Signals: {len(data['signals'])}개")
                print(f"  Default Function Index: {data['default_function']}")
                for sig_idx, sig_name in enumerate(data['signals']):
                    # sig_idx는 리스트의 인덱스, 실제 Function Index는 정렬된 순서
                    default_mark = " (DEFAULT)" if sig_idx == data['default_function'] else ""
                    print(f"    [Function Index {sig_idx}] {sig_name}{default_mark}")
            
            # 처음 3개 Signal 출력
            print("\n[Signal Map]")
            for idx, (sig_name, ball_locs) in enumerate(list(signal_map.items())[:3]):
                print(f"\nSignal: {sig_name}")
                print(f"  Used in: {', '.join(ball_locs)}")
        except Exception as e:
            print(f"오류: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("사용법: python excel_parser.py <엑셀파일경로>")
